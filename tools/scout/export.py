#!/usr/bin/env python3
"""Excel-файлы для курации сетов: файл на роль, 3 листа (Эконом/Комфорт/Премиум).
Запуск: python3 export.py [роль ...] из папки, где есть thumbs/ и xlsx/ (создать заранее)."""
import subprocess, re, io, os, sys, urllib.request, concurrent.futures as cf
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

PSQL=["docker","exec","-i","remlab-devdb","psql","-U","remlab","-d","remlab","-q","-v","ON_ERROR_STOP=1","-t","-A","-F","\x1f"]
def rows(q):
    r=subprocess.run(PSQL,input=q,capture_output=True,text=True)
    if r.returncode!=0: print(r.stderr[:400]); sys.exit(1)
    return [l.split('\x1f') for l in r.stdout.strip().split('\n') if l]

STOP=re.compile(r'\b(беж\w*|сер\w*|син\w*|зел[её]н\w*|коричн\w*|ч[её]рн\w*|бел\w*|графит\w*|латте|мокко|изумруд\w*|горчичн\w*|пудр\w*|роз\w*|голуб\w*|фиолет\w*|бордо\w*|венге|дуб\s?\w*|орех\w*|ясень|сонома|капучино|шоколад\w*|молочн\w*|крем\w*|песочн\w*|терракот\w*|оливк\w*|мятн\w*|лаванд\w*|карбон|антрацит|жемчужн\w*|сливов\w*|вельвет\w*|велюр\w*|шенилл\w*|рогожк\w*|экокож\w*|микровелюр\w*|правый|левый|угол|бархат\w*|тёмн\w*|темн\w*|светл\w*|глосс|люкс|найс|плюш\w*)\b', re.I)
def model_key(name):
    n=STOP.sub(' ', name.lower())
    n=re.sub(r'[^а-яa-z0-9 ]',' ',n); n=re.sub(r'\s+',' ',n).strip()
    return ' '.join(n.split()[:6])

def thumb(url, key):
    p=f"thumbs/{key}.png"
    if os.path.exists(p): return p
    try:
        small=url.replace('/big.jpg','/small.jpg').replace('/big.png','/small.png')
        if small.startswith('//'): small='https:'+small
        req=urllib.request.Request(small,headers={'User-Agent':'Mozilla/5.0'})
        data=urllib.request.urlopen(req,timeout=20).read()
        im=Image.open(io.BytesIO(data)).convert('RGB')
        im.thumbnail((110,82))
        im.save(p,'PNG')
        return p
    except Exception:
        return None

ROLES={
 'диван':'Диваны','кресло':'Кресла','пуф':'Пуфы-банкетки','столик':'Журнальные-столики',
 'тв-тумба':'ТВ-тумбы','стенка':'Стенки','стеллаж':'Стеллажи','полка':'Полки',
 'витрина':'Витрины-буфеты','комод':'Комоды','зеркало':'Зеркала',
 'стол обеденный':'Столы-обеденные','стул':'Стулья',
 'торшер':'Торшеры','лампа':'Настольные-лампы','люстра':'Люстры','бра':'Бра',
 'ковёр':'Ковры','камин':'Камины','ваза':'Вазы','статуэтка':'Статуэтки','часы':'Часы',
 'кашпо':'Кашпо','растение':'Растения','плед':'Пледы-покрывала','подушка':'Подушки','шторы':'Шторы'}
only=sys.argv[1:] or list(ROLES)
HDR=['Фото','Модель','Вариантов','Ш, см','Г, см','В, см','Площадь, м²','Цена от, ₽','Старая, ₽','Магазин','Бренд','Ссылка','ID']
TIERS=[('Эконом',0.10,0.40),('Комфорт',0.40,0.75),('Премиум',0.75,0.95)]
TYP_D={'диван':100,'кресло':90,'стенка':45,'стеллаж':35,'шкаф':58,'полка':25}

for role in only:
    fname=ROLES[role]
    data=rows(f"""select shop_mid, external_id, name, coalesce(brand,''), w_cm, d_cm, h_cm, len_cm, dia_cm,
      price_rub, coalesce(old_price_rub,0), shop, coalesce(image_url,''),
      urldecode.u from lr_roles,
      lateral (select replace(replace(replace(substring(url from 'goto=([^&]+)'),'%3A',':'),'%2F','/'),'%3F','?') u) urldecode
      where role='{role}' and price_rub is not null and image_url is not null and w_cm is not null
      order by price_rub;""")
    groups={}
    for r in data:
        k=(r[11], model_key(r[2]))
        g=groups.setdefault(k,{'rows':[],'minp':10**9})
        g['rows'].append(r); g['minp']=min(g['minp'],int(r[9]))
    reps=[]
    for (shop,mk),g in groups.items():
        rep=min(g['rows'], key=lambda r:int(r[9]))
        reps.append((rep,len(g['rows']),g['minp']))
    reps.sort(key=lambda x:x[2])
    if not reps: print(f"{fname}: пусто"); continue
    prices=[m for _,_,m in reps]
    def pct(p):
        i=max(0,min(len(prices)-1,int(p*len(prices))))
        return prices[i]
    wb=Workbook(); wb.remove(wb.active)
    total_rows=0
    for tname,lo,hi in TIERS:
        plo,phi=pct(lo),pct(hi)
        subset=[x for x in reps if plo<=x[2]<=phi]
        ws=wb.create_sheet(tname)
        ws.append(HDR)
        for c in ws[1]: c.font=Font(bold=True); c.fill=PatternFill('solid',fgColor='EEE5DC')
        ws.freeze_panes='A2'
        widths=[16,46,9,7,7,7,10,11,10,15,14,10,22]
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
        keys=[f"{r[0]}-{re.sub(r'[^A-Za-z0-9]','_',r[1])[:40]}" for r,_,_ in subset]
        with cf.ThreadPoolExecutor(8) as ex:
            thumbs=list(ex.map(lambda a: thumb(a[0][12],a[1]), zip((x[0] for x in subset),keys)))
        for idx,((r,nvar,minp),tp) in enumerate(zip(subset,thumbs),start=2):
            mid,eid,name,brand,w,d,h,ln,dia,price,old,shop,img,direct=r
            d_use=d or ln or ''
            approx=''
            if not d_use and role in TYP_D: d_use=TYP_D[role]; approx='≈'
            area=''
            try:
                if w and d_use: area=round(float(w)*float(d_use)/10000,2)
                elif dia: area=round(3.1416*(float(dia)/200)**2,2)
            except: pass
            ws.cell(idx,2,name).alignment=Alignment(wrap_text=True,vertical='top')
            ws.cell(idx,3,nvar)
            ws.cell(idx,4,float(w) if w else None)
            ws.cell(idx,5,(f"{approx}{d_use}" if approx else (float(d_use) if d_use else None)))
            ws.cell(idx,6,float(h) if h else None)
            ws.cell(idx,7,area or None)
            ws.cell(idx,8,int(price))
            if int(old)>int(price): ws.cell(idx,9,int(old))
            ws.cell(idx,10,shop); ws.cell(idx,11,brand or None)
            lc=ws.cell(idx,12,'открыть'); lc.hyperlink=direct; lc.font=Font(color='0563C1',underline='single')
            ws.cell(idx,13,f"{mid}:{eid}")
            ws.row_dimensions[idx].height=64
            if tp:
                im=XLImage(tp); im.anchor=f"A{idx}"; ws.add_image(im)
        total_rows+=len(subset)
    out=f"xlsx/{fname}.xlsx"; wb.save(out)
    print(f"{fname}: моделей {len(reps)}, строк в файле {total_rows} -> {out}", flush=True)
