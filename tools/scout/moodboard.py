#!/usr/bin/env python3
"""Мудборды сетов в Excel: один лист = один сет, крупные фото (240px из big), подписи, ссылки.
Читает sets.json (из compose.py). Выход: Сеты-гостиная.xlsx"""
import os, io, re, json, urllib.request, concurrent.futures as cf
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
RU={'neutral_light':'нейтр. светлый','neutral_grey':'нейтр. серый','neutral_dark':'нейтр. тёмный',
 'wood_light':'дерево светлое','wood_dark':'дерево тёмное','unknown':'не определён',
 'accent_terra':'акцент: терракота','accent_yellow':'акцент: жёлтый','accent_green':'акцент: зелёный',
 'accent_cyan':'акцент: бирюза','accent_blue':'акцент: синий','accent_violet':'акцент: фиолет',
 'accent_pink':'акцент: розовый','accent_red':'акцент: красный','accent_other':'акцент: прочий'}
from openpyxl.utils import get_column_letter
from PIL import Image

HERE=os.path.dirname(os.path.abspath(__file__))
BIG=os.path.join(HERE,'big'); os.makedirs(BIG,exist_ok=True)
import sys
SETS_F=sys.argv[1] if len(sys.argv)>1 else 'sets.json'
OUT_F=sys.argv[2] if len(sys.argv)>2 else 'Сеты-гостиная.xlsx'
sets=json.load(open(os.path.join(HERE,SETS_F)))
def big_thumb(url,mid,eid):
    p=os.path.join(BIG,f"{mid}-{re.sub(r'[^A-Za-z0-9]','_',eid)[:40]}.jpg")
    if os.path.exists(p): return p
    try:
        if url.startswith('//'): url='https:'+url
        req=urllib.request.Request(url,headers={'User-Agent':'Mozilla/5.0'})
        data=urllib.request.urlopen(req,timeout=25).read()
        im=Image.open(io.BytesIO(data)).convert('RGB'); im.thumbnail((240,190))
        im.save(p,'JPEG',quality=82)
        return p
    except Exception: return None

wb=Workbook(); wb.remove(wb.active)
COLS=4
for i,s in enumerate(sets,1):
    title=f"{i}. {s['band']} м² {s['tier']}"
    ws=wb.create_sheet(title[:31])
    ws.merge_cells('A1:H1')
    ws['A1']=f"Сет {i} — гостиная {s['band']} м², {s['tier'].upper()} · мебель на полу {s['fill_pct']}% · итого ≈ {s['total']:,} ₽".replace(',',' ')
    ws['A1'].font=Font(bold=True,size=14)
    items=list(s['items'].items())
    with cf.ThreadPoolExecutor(8) as ex:
        paths=list(ex.map(lambda kv: big_thumb(kv[1]['img'],kv[1]['mid'],kv[1]['eid']), items))
    for c in range(COLS):
        ws.column_dimensions[get_column_letter(c*2+1)].width=34
        ws.column_dimensions[get_column_letter(c*2+2)].width=4
    for idx,((role,it),p) in enumerate(zip(items,paths)):
        r0=3+(idx//COLS)*13; c0=(idx%COLS)*2+1
        col=get_column_letter(c0)
        ws.row_dimensions[r0].height=150
        if p:
            im=XLImage(p); im.anchor=f"{col}{r0}"; ws.add_image(im)
        q=f" ×{it['qty']}" if it.get('qty',1)>1 else ''
        dims=f"{it.get('w') or ''}×{it.get('d') or it.get('dia') or ''}"
        cell=ws.cell(r0+1,c0,f"{role.upper()}{q}: {it['name'][:60]}")
        cell.alignment=Alignment(wrap_text=True,vertical='top'); cell.font=Font(size=9)
        ws.row_dimensions[r0+1].height=26
        ws.cell(r0+2,c0,f"{dims} см · {it['price']:,} ₽ · {it['shop']}".replace(',',' ')).font=Font(size=9)
        lc=ws.cell(r0+3,c0,'открыть'); lc.hyperlink=it['url']; lc.font=Font(size=9,color='0563C1',underline='single')
        cc=ws.cell(r0+4,c0,f"цвет: {RU.get(it.get('cls','unknown'),it.get('cls',''))}")
        cc.font=Font(size=9)
        if it.get('rgb'):
            r,g,b=it['rgb']; hexc=f"{r:02X}{g:02X}{b:02X}"
            cc.fill=PatternFill('solid',fgColor=hexc)
            if (r*299+g*587+b*114)/1000<128: cc.font=Font(size=9,color='FFFFFF')
        ws.cell(r0+5,c0,f"{it['mid']}:{it['eid']}").font=Font(size=7,color='AAAAAA')
out=os.path.join(HERE,OUT_F); wb.save(out)
print('OK:',out,f'({len(sets)} листов)')
