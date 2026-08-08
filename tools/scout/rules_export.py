#!/usr/bin/env python3
"""Экспорт ВСЕГО действующего свода правил в один .xlsx — для ревизии владельцем.

Листы: оглавление · канон occupancy.json (один файл на оба движка) · zones.json (группы/зоны/
inventory-prior/иерархия) · weights.json (веса скоринга) · composition.json (доли пола по
метражу) · size-bands.json (размерные priors, не гейты) · proportions.json (hard/soft) ·
коды валидатора (из validate.py, с severity и местом).

Запуск: ~/venvs/scout/bin/python rules_export.py [выходной .xlsx]
"""
import datetime
import json
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
SOLVER = os.path.join(HERE, '..', '..', 'services', 'planner-solver')
_pos = [a for a in sys.argv[1:] if not a.startswith('--')]
# --date=YYYY-MM-DD: дата пакета, если «сегодня» владельца/рефери не совпадает с UTC машины
# (рефери-финал 08.08 P0.1: имя файла и оглавление обязаны совпадать — оба берутся отсюда)
_dt = next((a.split('=', 1)[1] for a in sys.argv[1:] if a.startswith('--date=')), None)
STAMP = _dt or datetime.date.today().isoformat()
OUT = _pos[0] if _pos else os.path.expanduser(
    f'~/rules-{"for-referee" if "--referee" in sys.argv else "svod"}-{STAMP}.xlsx')

FILES = [
    ('occupancy.json', os.path.join(SOLVER, 'rules', 'occupancy.json'),
     'КАНОН расстановки: шкалы расстояний по метражу, доли, ярусы, именованные правила '
     '(один файл на оба движка; tools/scout/occupancy.json — синхронизированная копия)'),
    ('zones.json', os.path.join(SOLVER, 'rules', 'zones.json'),
     'Зоны-first (ADR-0074): 10 посадочных групп с футпринтами, внутренняя схема группы, '
     'функциональные зоны, inventory-prior по usable-площади, лексикографическая иерархия'),
    ('weights.json', os.path.join(SOLVER, 'rules', 'weights.json'),
     'Веса мягких термов скоринга солвера'),
    ('composition.json', os.path.join(HERE, 'composition.json'),
     'Состав сетов: доли площади пола по ролям в каждом метраже'),
    ('size-bands.json', os.path.join(HERE, 'size-bands.json'),
     'Размерные priors/preferred bands: in-band приоритет, вне диапазона штраф — НЕ отсев (W3)'),
    ('proportions.json', os.path.join(HERE, 'proportions.json'),
     'Пропорции: hard functional filters + soft aesthetic preferences (флаг hard, 5.1)'),
]


def flatten(obj, path=''):
    rows = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            rows += flatten(v, f'{path}.{k}' if path else str(k))
    elif isinstance(obj, list) and any(isinstance(x, (dict, list)) for x in obj):
        for i, v in enumerate(obj):
            rows += flatten(v, f'{path}[{i}]')
    else:
        val = json.dumps(obj, ensure_ascii=False) if isinstance(obj, list) else obj
        rows.append((path, val))
    return rows


def add_sheet(wb, title, rows, headers):
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append(list(r))
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            widths[c.column] = min(90, max(widths.get(c.column, 10),
                                           len(str(c.value or '')) + 2))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'
    return ws


def validator_codes():
    """Каждый вызов _v(...) разбираем ЦЕЛИКОМ (многострочно): severity часто на следующей
    строке, и однострочный парс оставлял «Жёсткость» пустой — внешний аудит 08.08 счёл это
    отсутствием источника истины. Дефолт _v без Severity — HARD (см. сигнатуру _v)."""
    src = open(os.path.join(SOLVER, 'planner', 'validate.py')).read()
    lines = src.splitlines()
    funcs = []
    for i, line in enumerate(lines, 1):
        m = re.match(r'def (check_\w+)', line)
        if m:
            rest = src.split(line, 1)[1]
            d = re.match(r'[^"]*"""(.*?)"""', rest, re.S)
            funcs.append((i, m.group(1), ' '.join(d.group(1).split())[:200] if d else ''))
    def func_of(ln):
        cur = ('', '')
        for fi, fn, doc in funcs:
            if fi <= ln:
                cur = (fn, doc)
        return cur
    seen = {}
    for m in re.finditer(r'_v\(\s*"([A-Z][A-Z_]{3,})"', src):
        ln = src[:m.start()].count('\n') + 1
        depth = 0
        j = m.start()
        while j < len(src):
            if src[j] == '(':
                depth += 1
            elif src[j] == ')':
                depth -= 1
                if depth == 0:
                    break
            j += 1
        call = src[m.start():j + 1]
        code = m.group(1)
        sev = ('SOFT' if 'Severity.SOFT' in call else 'HARD')
        fn, doc = func_of(ln)
        if code not in seen:
            seen[code] = [code, sev, fn, f'validate.py:{ln}', doc]
        elif sev == 'SOFT' and seen[code][1] == 'HARD':
            seen[code][1] = 'HARD/SOFT (по условию)'
    return sorted(seen.values())


# --- Режим «для рефери» (--referee): позиция по аудиту + спорные вопросы -------------------
AUDIT_RESPONSE = [
 ('P0.1 Жёсткость кодов пуста', 'ИСПРАВЛЕНО 08.08', 'Дефект экспортёра (severity на соседней строке не парсилась). Теперь: rules/severity.json — реестр классов H0/H1/S1/S2 на каждый код + тест test_severity_registry, механически сверяющий реестр с кодом валидатора. В листе «коды валидатора» — фактическая жёсткость и класс.'),
 ('P0.2 CHAIR_ORPHAN 40 vs 100 см', 'ИСПРАВЛЕНО (док), КОД БЫЛ ВЕРЕН', 'В коде всегда 40 см и ТОЛЬКО для роли «стул» (обеденный); кресло — отдельная роль, правило его не касается. «100 см» — устаревший docstring, исправлен. Subtype-разделение, которого требовал аудит, уже было по построению ролей.'),
 ('P0.3 Столик от площади', 'ПРИНЯТО, W2 РЕАЛИЗОВАНО 08.08', 'Валидатор: hard 32–50 (tight-fallback), комфорт 36–46 soft (SOFA_TABLE_COMFORT), от площади НЕ зависит. Area-шкала dynamic.sofa_table_cm помечена legacy. Скоринг переведён на те же вилки.'),
 ('P0.4 ТВ-дистанция от площади', 'ПРИНЯТО, W2 РЕАЛИЗОВАНО 08.08', 'Primary — диагональ: ТВ в сете нет (рисует генератор), диагональ ОЦЕНИВАЕТСЯ по ширине тумбы (экран≈70% тумбы, /0.872 для 16:9). Вилка 1.2–2.5 диагонали hard, >2.0 soft. Area-шкала — только фолбэк при неизвестной ширине. ВОПРОС РЕФЕРИ №2: корректна ли оценка диагонали по тумбе.'),
 ('P0.5 Несколько floor-cap', 'ЧАСТИЧНО; РАЗВОД СУЩНОСТЕЙ В ОЧЕРЕДИ (W7 IR)', 'Фактически: floor_cap_pct (dynamic, по метражу) — операционный кап солвера/состава; composition.global_floor_cap — legacy-фолбэк. Развод на 4 метрики (physical/usable/visual/zone) — часть будущего constraint-IR.'),
 ('P0.6 corner_sofa_must_be_in_corner', 'УЖЕ БЫЛО КАК РЕКОМЕНДОВАНО', 'CORNER_SOFA_ADRIFT — SOFT (класс S2), floating разрешён: в комнатах 50+ Г-диван ОБЯЗАН отплывать (комментарий в коде validate.py). Флаг в occupancy — предпочтение кандидатов, не hard.'),
 ('P0.7 tv_not_on_window_wall hard ban', 'УЖЕ БЫЛО КАК РЕКОМЕНДОВАНО', 'TV_ON_WINDOW_WALL — SOFT (S1), штраф за блики, не запрет. Раздельные проверки window-obstruction (WINDOW_BLOCKED, hard) и glare (этот SOFT) уже есть.'),
 ('P0.8 base=диван+столик+ТВ конфликтует с zones-first', 'ПРИНЯТО, РЕАЛИЗОВАНО Z4', 'Состав идёт scenario→usable→группа→спутники (compose2+zones.json). placement_tiers.base — ярусы ПРИОРИТЕТА размещения у солвера (что терять последним), не обязательный состав; столик в ≤8 м² band опционален.'),
 ('P0.9 size-bands hard gate', 'ПРИНЯТО, W3 РЕАЛИЗОВАНО 08.08', 'Вилка → приоритет: in-band первыми + бонус, вне — штраф −1.5, НЕ выбывание. Жёстко отсеивают только функция/пропорции/качество карточки; физическую невозможность ловит солвер геометрией.'),
 ('P0.10 SOFA_AIM 30° без сценария', 'ПРИНЯТО ЧАСТИЧНО: H1-класс', 'Класс H1 = «жёстко в ТЕКУЩЕМ продукте (media-гостиная — единственный сценарий)». При добавлении сценариев конфиг перейдёт в условный. Снимать сейчас — нет продукта-потребителя.'),
 ('P1 5.1 chair_requires_dining_table', 'УЖЕ БЫЛО', 'Только роль «стул»; кресло/акцентное — отдельные роли.'),
 ('P1 5.3-5.4 rug-правила', 'УЖЕ БЫЛО', 'Две легальные схемы (front-legs / под столиком с выносами) + вариант «вся мебель на большом ковре» (правка владельца). Схема «только под столиком» — фолбэк, не приоритет.'),
 ('F2S gaps (articulation/reach/workflow/balance)', 'ЧАСТИЧНО / СКОУП', 'Articulation прямоугольной аппроксимацией УЖЕ есть (фасады 50–80, ящик 76, ноги 46 — clearances.py). pair_symmetry добавлен (W5, ярус S2). Reach/posture/workflow/acoustics — сознательно вне скоупа продукта (см. страницу-обоснование).'),
 ('DFS/beam «слабее»', 'ФОРМУЛИРОВКИ ИСПРАВЛЕНЫ', 'Только «на нашем бенчмарке»; страница-обоснование переписана, оба движка сравнимы в лоб: solver_run --engine beam|zoned|dfs.'),
 ('3D-FRONT «реальные»', 'ИСПРАВЛЕНО', 'Везде: «профессионально спроектированные СИНТЕТИЧЕСКИЕ сцены», подмножество фильтровали сами (5 742 гостиных).'),
 ('Provenance без SHA', 'ПРИНЯТО, W4 РЕАЛИЗОВАНО 08.08', 'infinigen@25a7d28 src/infinigen_examples/constraints/home.py (секция livingrooms); Holodeck@362b8ed ai2holodeck/generation/small_objects.py; помечено derived, дата сверки.'),
]
REFEREE_QUESTIONS = [
 ('Q1', 'Обеденная группа: гейт «остаток usable ≥ 6 м²» (после вычета футпринта посадочной группы). Порог взят инженерно (стол 120×75 + стулья + отступы). Верна ли величина и сам подход?'),
 ('Q2', 'ТВ-дистанция: диагональ ОЦЕНИВАЕТСЯ по ширине тумбы (экран ≈ 70% ширины тумбы, диагональ = ширина экрана / 0.872). ТВ как товар в сете отсутствует. Допустима ли такая оценка как primary?'),
 ('Q3', 'DEAD_ZONE_BEHIND_SOFA: полоса за спинкой дивана на ВСЮ ширину комнаты (не только тень дивана) — вердикт владельца после сета с камином в углу за спиной. Не слишком ли широко для больших комнат?'),
 ('Q4', 'NOT_AT_WALL (корпусная мебель спинкой к стене ≤20 см) — класс H1. Room-divider сценарии (стеллаж как перегородка) пока запрещены. Оставить до появления сценария или ослабить сейчас?'),
 ('Q5', 'floor_cap_pct по метражу (40–50% малые → 26–34% в 50+) — операционный кап. Смешивает physical footprint и visual density (замечание P0.5). Достаточно ли до IR-рефактора?'),
 ('Q6', 'Камин: планируем FIREPLACE_FAR_FROM_SEATING (S1) при >450 см до посадки + требование видимости с главной посадки (сектор ≤~75° от оси взгляда). Согласен с порогами 200–450 и углом?'),
 ('Q7', 'Диван спинкой к окну: enforce отступ 15–20 см + «спинка не выше низа стекла» (правило есть в данных, в валидатор не доведено). Какой класс — H1 или S1?'),
]

# Финальная строка приёмки: читаем свежий jsonl, если он полон; иначе честная пометка
def _acceptance_final() -> str:
    import json as _json
    p = os.path.join(HERE, 'acceptance-report-zoned.jsonl')
    try:
        rows = [_json.loads(l) for l in open(p)]
    except OSError:
        return 'прогон не найден на момент экспорта'
    ok = sum(1 for r in rows if r.get('ok'))
    note = '' if len(rows) >= 252 else f' (ПРОМЕЖУТОЧНО, прогон идёт: {len(rows)}/252 сцен)'
    import collections as _c
    fails = _c.Counter()
    for r in rows:
        if not r.get('ok'):
            for f in (r.get('fails') or ['unplaced base']):
                fails[str(f).replace('FAIL', '').strip()[:40]] += 1
    top = '; '.join(f'{k}×{v}' for k, v in fails.most_common(6)) or 'провалов нет'
    return f'{ok}/{len(rows)} чистых{note}; провалы: {top}'


ACCEPTANCE_FINAL = _acceptance_final()

# Детальный отчёт внедрения 08.08: пункт рефери → изменение → якорь в коде → проверка
IMPLEMENTATION_REPORT = [
 ('Q5 floor-cap не HARD', 'FLOOR_OVERFILL: Severity HARD→SOFT; класс H1→S1 в реестре',
  'planner/validate.py check_floor_cap; rules/severity.json',
  'test_floor_overfill_is_soft; test_severity_registry (механическая сверка реестр↔код)'),
 ('Q6 камин S1-prior', 'НОВЫЙ чек FIREPLACE_FAR_FROM_SEATING (S1): дистанция вне 200–450 см до посадки ИЛИ вне сектора видимости 75° от оси взгляда дивана; пороги из zones.json fireplace.distance_to_seating_cm',
  'planner/validate.py check_fireplace_seating', 'test_fireplace_far_is_soft (far→S1, в вилке+в поле→чисто)'),
 ('Q7 окно раздельно', 'Доступ был H0 (WINDOW_BLOCKED/RADIATOR); НОВЫЕ: SOFA_WINDOW_GAP (S1, спинка к окну ближе 15 см, из window_sofa.min_offset) и SOFA_BACK_ABOVE_SILL (S2, спинка выше sill_cm проёма)',
  'planner/validate.py check_window_sofa', 'test_sofa_window_gap_and_sill'),
 ('Q1/3.3 ярусы = retention-priority', 'Не влезшие dining/storage/optional дропаются ЯРУСОМ в skipped_optional (не провал сцены); обеденная группа атомарна: стол + ≥2 стульев или ничего; 6 м² остаётся префильтром состава; дроп виден наружу (SKIPPED в выводе солвера и поле skipped приёмки — no silent caps)',
  'planner/beam.py solve; tools/scout/solver_run.py; acceptance_run.py', 'test_dining_storage_drop_not_fail'),
 ('Q3 локализация полосы', '_behind_strip: ширина дивана + 100 см бокового запаса (не вся комната); камин исключён из DEAD_BEHIND_ROLES (focal-behind — угловым чеком)',
  'planner/validate.py _behind_strip, DEAD_BEHIND_ROLES', 'test_verdicts_0807 (камин→угловой чек; торшер за плечом→полоса)'),
 ('Q4 subtype-флаги', 'requires_wall_back=[тв-тумба,шкаф,комод,стенка,витрина,стеллаж,камин]; room_divider_capable=[стеллаж]; room_divider_capable_active=[] (активация без переписывания severity)',
  'rules/occupancy.json layout_rules; validate.py _wall_only_roles', 'поведение текущего скоупа без изменений (79 тестов)'),
 ('Q2 ТВ от дистанции', 'Валидатор: вилка от существования диагонали при приоре 0.70–0.90 экрана к тумбе; генератор: distance-first (диагональ ≈ дистанция/1.6 по FOV ~30°, clamp 70–90% тумбы)',
  'validate.py check_distances; tools/scout/viz_final.py zones_brief', 'фикстуры manual_layout валидны в новой вилке'),
 ('5.1 эстетика ≠ отсев', 'Флаг hard у каждого правила пропорций: 6 эстетических (chair_h, rug×2, pouf_area, sofa_vs_wall, storage_vs_wall) вне allowed → штраф −1.5, товар НЕ выбывает; hard остались: длина/высота столика, высота пуфа, тумба-не-уже-ТВ',
  'tools/scout/proportions.json (hard); proportions.py', 'подбор: физику проверяет геометрия солвера'),
 ('5.2 один канон высоты столика', 'Канон: zones height_vs_seat_cm [-5,0] (вровень или до 5 см ниже сиденья); proportions table_h_vs_seat — производная (allowed 0.78–1.08 — выше сиденья убрано, preferred 0.89–1.0)',
  'rules/zones.json (_canon_height); proportions.json', '—'),
 ('5.3 классы маршрутов', 'Авторитетная карта: primary_route=passage_main[90,107]; secondary_route=walkway[76,91]; object_access=gaps 60–75; tight_fallback=[46,61]; переименование ключей — в constraint-IR (W7)',
  'rules/occupancy.json distances_cm._route_classes', '—'),
 ('5.4 narrow-room = кандидаты', 'dynamic.narrow_room помечен как кандидат-шаблоны генерации, НЕ канон-констрейнты (enforcement в валидаторе и не было); вытянутые — солверная работа (очередь №1)',
  'rules/occupancy.json dynamic.narrow_room._note', '—'),
 ('P0.8 authoritative floor-cap', 'Авторитетный метрик — band-scale floor_cap_pct валидатора; composition.global_floor_cap — legacy-фолбэк подбора; развод на 4 метрики — W7',
  'validate.py check_floor_cap', '—'),
 ('Гигиена A/B (наша находка)', 'Под ENGINE=zoned после планнера гонялись 6 DFS-попыток и при меньшем числе нарушений подменяли результат (контаминация A/B + таймауты) — убрано: zoned = чистый планнер',
  'tools/scout/solver_run.py', 'таймауты ушли, скорость прогона ×2'),
]

# Вердикты рефери 08.08 и что сделано по ним (арбитраж завершён)
REFEREE_VERDICTS = [
 ('Q1', '6 м² — только permissive-префильтр; финал по фактической геометрии', 'ПРИНЯТО, так и работает: гейт в составе — префильтр, финальное слово за солвером (клиренсы стульев/проходы). Дополнено 08.08: не влезшие dining/storage дропаются ЯРУСОМ в skipped (retention-priority, п.3.3 рефери), а не валят сцену.'),
 ('Q2', 'stand→TV 70% — только fallback; primary — от sofa-distance/FOV (~30°)', 'ПРИНЯТО, СДЕЛАНО 08.08: валидатор принимает дистанцию, под которую СУЩЕСТВУЕТ диагональ при приоре экран/тумба 0.70–0.90 (hard [1.2·d_min, 2.5·d_max], не точка 0.70); генератору передана distance-first инструкция (диагональ ≈ дистанция/1.6, clamp 70–90% тумбы, не шире тумбы). Калибровка приора по каталогу — при накоплении рендеров.'),
 ('Q3', 'room-wide полоса слишком широка; локализовать + отдельный focal-behind', 'ПРИНЯТО, СДЕЛАНО 08.08 (решение владельца: «делаем как рефери»): полоса локализована — ширина дивана + 100 см бокового запаса (кейс «торшер за плечом» покрыт); камин исключён из полосы — focal-behind ловит угловой чек FIREPLACE_FAR_FROM_SEATING (сектор 75°). Отдельный FOCAL_BEHIND_MAIN_SEAT — при мульти-focal сценариях.'),
 ('Q4', 'H1 допустим в текущем скоупе, но только с subtype-привязкой; open shelving divider — exemption', 'ПРИНЯТО, СДЕЛАНО 08.08: requires_wall_back / room_divider_capable (стеллаж) / room_divider_capable_active в occupancy.layout_rules; валидатор читает из данных. Активация divider-exemption — переносом роли в _active, severity переписывать не нужно (сейчас пусто: divider-сценариев в продукте нет).'),
 ('Q5', 'dynamic floor-cap — временно ок, но НЕ HARD; один authoritative metric', 'ПРИНЯТО, СДЕЛАНО 08.08: FLOOR_OVERFILL → S1 (SOFT). Authoritative — band-scale floor_cap_pct валидатора; composition.global_floor_cap — legacy-фолбэк подбора; развод на 4 метрики — в IR (W7).'),
 ('Q6', '200–450 и ≤75° — только S1-prior, safety отдельно от композиции', 'ПРИНЯТО, СДЕЛАНО 08.08: FIREPLACE_FAR_FROM_SEATING (S1) — вне вилки 200–450 ИЛИ вне сектора ≤75°; safety-clearance прибора — вне скоупа (метаданные товара, TODO). Тест в test_zones.'),
 ('Q7', 'разделить: доступ H0/H1; зазор S1 (preferred 25–30, tight 15–20); спинка выше стекла S2', 'ПРИНЯТО, СДЕЛАНО 08.08: доступ/радиатор уже H0 (WINDOW_BLOCKED/RADIATOR); SOFA_WINDOW_GAP (S1, tight-минимум 15); SOFA_BACK_ABOVE_SILL (S2, порог — sill_cm проёма). Preferred 25–30 — в данные при калибровке.'),
]

def referee_sheets(wb):
    ws = add_sheet(wb, 'ответ на аудит', AUDIT_RESPONSE,
                   ['Пункт аудита', 'Статус', 'Позиция проекта / что сделано'])
    ws.column_dimensions['C'].width = 110
    ws2 = add_sheet(wb, 'вопросы рефери', REFEREE_QUESTIONS, ['№', 'Вопрос'])
    ws2.column_dimensions['B'].width = 120
    accpt = [
     ('Приёмочный набор', '252 зафиксированные сцены: 126 базовых прямоугольных + 63 вытянутых 1:1.5 + 63 сложных контура (эркер/пилоны/трапеция), acceptance-scenes.json в git'),
     ('Старый движок (beam), новые составы', '119/252 чистых (47%); медиана мягкого балла 11.1'),
     ('Зонный движок, новые составы', '239/252 чистых (95%); медиана 9.7; сцен хуже старого — 0'),
     ('Зонный по типам сцен', 'база 124/126, вытянутые 56/63, эркер 20/21, пилоны 21/21, трапеция 18/21'),
     ('Решение', 'зонный — боевой дефолт с 08.08; beam остаётся для A/B (--engine beam)'),
     ('Прогон с W-правками (08.08, до пакета рефери)', '221/252 при СТАРОМ понимании required; разбор: 22 из 31 провала — не влезшие dining/storage в новых амбициозных составах (закрыто дропом ярусом по Q1/3.3), 2 таймаута — DFS-фолбэк под лейблом zoned (убран), остальное — диван↔ТВ/sliver'),
     ('Финальный прогон с ПОЛНЫМ пакетом рефери (партии 1+2)', ACCEPTANCE_FINAL),
    ]
    ws3 = add_sheet(wb, 'приёмка (контекст)', accpt, ['Что', 'Результат'])
    ws3.column_dimensions['B'].width = 120
    ws4 = add_sheet(wb, 'вердикты рефери', REFEREE_VERDICTS,
                    ['№', 'Вердикт рефери (08.08)', 'Решение проекта / статус'])
    ws4.column_dimensions['B'].width = 70
    ws4.column_dimensions['C'].width = 110
    ws5 = add_sheet(wb, 'внедрение (детально)', IMPLEMENTATION_REPORT,
                    ['Пункт', 'Изменение', 'Якорь в коде', 'Проверка'])
    ws5.column_dimensions['B'].width = 90
    ws5.column_dimensions['C'].width = 55
    ws5.column_dimensions['D'].width = 55


def export_json():
    """--json: тот же пакет одним машиночитаемым файлом (для внешнего ИИ; вопрос владельца
    08.08 — JSON первичен). Файлы правил кладём СЫРЬЁМ (они и есть источник истины)."""
    sev = json.load(open(os.path.join(SOLVER, 'rules', 'severity.json')))
    doc = {
        '_meta': {
            'project': 'remlab layout rule pack', 'date': STAMP,
            'what': ('полный свод правил расстановки: сырые файлы правил + реестр severity '
                     '(H0/H1/S1/S2) + 48 кодов валидатора с местом в коде + арбитраж рефери '
                     '(вопросы, вердикты, внедрение) + контекст приёмки 252 сцен'),
        },
        'rule_files': {name: json.load(open(path)) for name, path, _ in FILES},
        'severity_registry': sev,
        'validator_codes': [
            {'code': c, 'hardness': s, 'class': sev['codes'].get(c),
             'function': fn, 'where': loc, 'doc': d}
            for c, s, fn, loc, d in validator_codes()],
        'audit_response': [dict(zip(('point', 'status', 'position'), r)) for r in AUDIT_RESPONSE],
        'referee_questions': [dict(zip(('q', 'question'), r)) for r in REFEREE_QUESTIONS],
        'referee_verdicts': [dict(zip(('q', 'verdict', 'resolution'), r)) for r in REFEREE_VERDICTS],
        'implementation_report': [dict(zip(('point', 'change', 'code_anchor', 'verified_by'), r))
                                  for r in IMPLEMENTATION_REPORT],
        'acceptance': {'final_run': ACCEPTANCE_FINAL,
                       'history': {'old_beam': '119/252 (47%)',
                                   'zoned_pre_referee': '239/252 (95%)',
                                   'zoned_referee_package': '243/252 (96.4%)',
                                   'zoned_full_with_dining': '238/252 (94.4%)'}},
    }
    out = os.path.splitext(OUT)[0] + '.json'
    json.dump(doc, open(out, 'w'), ensure_ascii=False, indent=1)
    print(f'{out}: {len(doc["rule_files"])} файлов правил, {len(doc["validator_codes"])} кодов')
    return out


def main():
    if '--json' in sys.argv:
        export_json()
        return
    wb = Workbook()
    ws = wb.active
    ws.title = 'Оглавление'
    ws.append(['Свод правил remlab — полный экспорт', ''])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'Дата: {STAMP}', ''])
    ws.append(['Лист', 'Что это'])
    for name, _, note in FILES:
        ws.append([name, note])
    ws.append(['коды валидатора', 'все именованные проверки размещения: код, жёсткость, '
               'функция и строка в validate.py'])
    for col, w in (('A', 28), ('B', 110)):
        ws.column_dimensions[col].width = w
    for name, path, _ in FILES:
        data = json.load(open(path))
        add_sheet(wb, name, flatten(data), ['Правило (путь в файле)', 'Значение'])
    sev_reg = json.load(open(os.path.join(SOLVER, 'rules', 'severity.json')))
    rows = [r + [sev_reg['codes'].get(r[0], '')] for r in validator_codes()]
    add_sheet(wb, 'коды валидатора', rows,
              ['Код', 'Жёсткость (код)', 'Проверка', 'Место', 'Описание проверки',
               'Класс (severity.json: H0 физика / H1 функция сценария / S1 эргономика / S2 эстетика)'])
    if '--referee' in sys.argv:
        referee_sheets(wb)
    wb.save(OUT)
    print(f'{OUT}: {len(wb.sheetnames)} листов')


if __name__ == '__main__':
    main()
