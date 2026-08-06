#!/usr/bin/env python3
"""Выгрузка таблицы «категория × признак → стиль» в Excel — для владельца.

Показывает ровно то, по чему считается стиль: какой признак у какого типа товара спрашивается,
и что каждое его значение даёт каждому из шести стилей — в плюс или в минус, и насколько.

  ~/venvs/scout/bin/python matrix_xlsx.py
"""
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

OUT = os.path.expanduser('~/scout-backups/style-matrix.xlsx')
STYLES = ['сканди', 'современный', 'минимализм', 'лофт', 'неоклассика', 'джапанди']
W = {'маркер': 3.0, 'поддержка': 1.5, 'фон': 0.6}
GREEN = PatternFill('solid', fgColor='DCEDDA')
STRONG = PatternFill('solid', fgColor='A8D5A2')
RED = PatternFill('solid', fgColor='F6D5D2')
STRONG_RED = PatternFill('solid', fgColor='EBA9A3')
HEAD = PatternFill('solid', fgColor='2F4858')
GREY = PatternFill('solid', fgColor='EFEDE8')


def cell_text(t: dict) -> tuple[str, object]:
    w = W[t['tier']] * t.get('sign', 1)
    txt = f'{t["tier"]} {w:+.1f}'
    if w >= 3:
        return txt, STRONG
    if w > 0:
        return txt, GREEN
    if w <= -3:
        return txt, STRONG_RED
    return txt, RED


def main() -> None:
    matrix = json.load(open(os.path.join(HERE, 'style-matrix.json')))
    wb = Workbook()

    ws = wb.active
    ws.title = 'Признаки по типам'
    head = ['Категория', 'Группа', 'Что спрашиваем', 'Ответ', 'Свой маркер'] + STYLES + ['Вето']
    ws.append(head)
    for i, _ in enumerate(head, 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = HEAD
        c.alignment = Alignment(vertical='center', wrap_text=True)
    row = 2
    for role, m in sorted(matrix.items()):
        for attr, a in m['attrs'].items():
            for val, cellv in a['values'].items():
                ws.cell(row=row, column=1, value=role)
                ws.cell(row=row, column=2, value=m['group'])
                ws.cell(row=row, column=3, value=a['q'])
                ws.cell(row=row, column=4, value=val.replace('_', ' '))
                ws.cell(row=row, column=5, value='да' if a['own_marker'] else '')
                for j, st in enumerate(STYLES, 6):
                    t = (cellv['tiers'] or {}).get(st)
                    if not t:
                        continue
                    txt, fill = cell_text(t)
                    c = ws.cell(row=row, column=j, value=txt)
                    c.fill = fill
                ws.cell(row=row, column=12, value=', '.join(cellv['veto']) if cellv['veto'] else '')
                if row % 2 == 0:
                    for j in (1, 2, 3, 4, 5):
                        ws.cell(row=row, column=j).fill = GREY
                row += 1
    widths = [16, 11, 22, 24, 11] + [17] * 6 + [26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'E2'
    ws.auto_filter.ref = f'A1:L{row - 1}'

    # лист 2: какие вопросы задаются каждому типу товара
    ws2 = wb.create_sheet('Вопросы по типам')
    ws2.append(['Категория', 'Группа', 'Сколько вопросов', 'Вопросы', 'Свои маркеры'])
    for i in range(1, 6):
        c = ws2.cell(row=1, column=i)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = HEAD
    r = 2
    for role, m in sorted(matrix.items()):
        qs = [a['q'] for a in m['attrs'].values()]
        own = [a['q'] for a in m['attrs'].values() if a['own_marker']]
        ws2.cell(row=r, column=1, value=role)
        ws2.cell(row=r, column=2, value=m['group'])
        ws2.cell(row=r, column=3, value=len(qs))
        ws2.cell(row=r, column=4, value=', '.join(qs))
        ws2.cell(row=r, column=5, value=', '.join(own))
        r += 1
    for i, w in enumerate([16, 11, 16, 74, 44], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # лист 3: как это считается
    ws3 = wb.create_sheet('Как считается')
    lines = [
        ('Шаг', 'Что происходит'),
        ('1. Категория', 'Роль товара берётся из дерева категорий фида (products.cat_role), '
                         'а не из слов в названии: карниз не штора, садовый вазон не кашпо.'),
        ('2. Вопросы', 'Модели отправляются вопросы ИМЕННО ЭТОГО типа товара — лист «Вопросы по '
                       'типам». У пледа спрашиваем плетение, край и узор; у люстры каркас, плафон, '
                       'материал плафона и отделку металла. Модель отвечает, ЧТО ВИДИТ, '
                       'а не «какой это стиль».'),
        ('3. Баллы', 'Каждый ответ даёт баллы стилям по рангу: маркер +3.0 (увидел — почти '
                     'наверняка этот стиль), поддержка +1.5 (согласуется), фон +0.6 (намёк). '
                     'Противоречащий признак даёт столько же со знаком минус.'),
        ('4. Вето', 'Некоторые ответы делают стиль невозможным: хрусталь и свечной канделябр '
                    'убивают лофт и минимализм, открытая лампа Эдисона — неоклассику, '
                    'офисная крестовина на колёсах — всё домашнее.'),
        ('5. Редкость', 'Ранг понижается, если признак частый ИМЕННО В ЭТОЙ категории: маркер '
                        'обязан быть редким. Тонкая металлическая опора у столика обычна, '
                        'у дивана — редкость, и весит она по-разному.'),
        ('6. Достаточность', 'Стиль не может выиграть без единого положительного маркера — иначе '
                             'он побеждает на том, чего у вещи НЕТ (декора нет, линии прямые), '
                             'а это верно для любой категории.'),
        ('7. Итог', 'Баллы шести стилей сравниваются между собой внутри одного товара и '
                    'переводятся в шкалу 0–10. Оценка умножается на уверенность: чем меньше '
                    'признаков модель разглядела, тем ближе результат к нейтральной пятёрке.'),
        ('', ''),
        ('Проверено', 'Прямая классификация стиля по одному предмету даёт 0.41–0.49 точности '
                      '(Bonn Furniture Styles, 90 298 фото, 17 стилей) — поэтому спрашиваем '
                      'признаки, а стиль собираем правилами.'),
        ('Источники', 'chairish и britannica — неоклассика; 2modern и AllModern — сканди; '
                      'salterspiralstair и learncalifornia — лофт; metercube — минимализм '
                      'против современного; moderncre8ve — джапанди; kathykuohome и luxdeco — '
                      'подлокотники и ножки; rockymountainhardware — ручки как ключ к стилю '
                      'корпусной мебели; arteriorshome — типы люстр; wellwoven и nazmiyal — ковры.'),
    ]
    for i, (a, b) in enumerate(lines, 1):
        ws3.cell(row=i, column=1, value=a).font = Font(bold=True)
        ws3.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical='top')
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 116
    for i in range(1, len(lines) + 1):
        ws3.row_dimensions[i].height = 46

    wb.save(OUT)
    print(f'таблица: {OUT}')
    print(f'строк с признаками: {row - 2}, категорий: {len(matrix)}')


if __name__ == '__main__':
    main()
